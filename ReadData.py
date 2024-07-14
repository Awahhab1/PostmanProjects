__author__ = 'abdul.wahhab'

# Helper file to read data from files
import csv
import xlrd
import datetime
import unittest
import xlsxwriter
global worksheet , workbook
from openpyxl import Workbook

def read_order_file():

    workbook =xlrd.open_workbook("InputFiles/OrderDetails.xlsx")
    worksheet=workbook.sheet_by_index(0)
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    table=list()
    record=list()

    for x in range(total_rows):
        for y in range(total_columns):


            record.append(worksheet.cell(x,y).value)
        if x !=0:
            table.append(record)

        record=[]
    return table


#def read_campaign_data():



def read_campaign_file():
    workbook =xlrd.open_workbook("InputFiles/CampaignDetails.xlsx")
    worksheet=workbook.sheet_by_index(0)
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    table=list()
    record=list()

    for x in range(total_rows):
        for y in range(total_columns):


            record.append(worksheet.cell(x,y).value)
        if x !=0:
            table.append(record)

        record=[]
    #print(table)
    return table


def read_application_file():
    workbook =xlrd.open_workbook("InputFiles/ApplicationDetails.xlsx")
    worksheet=workbook.sheet_by_index(0)
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    table=list()
    record=list()

    for x in range(total_rows):
        for y in range(total_columns):


            record.append(worksheet.cell(x,y).value)
        if x !=0:
            table.append(record)

        record=[]
    #print(table)
    return table


def read_adunit_file():
    workbook =xlrd.open_workbook("InputFiles/AdunitDetails.xlsx")
    worksheet=workbook.sheet_by_index(0)
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    table=list()
    record=list()

    for x in range(total_rows):
        for y in range(total_columns):


            record.append(worksheet.cell(x,y).value)
        if x !=0:
            table.append(record)

        record=[]
    #print(table)
    return table

def read_targeting_file():
    workbook =xlrd.open_workbook("InputFiles/Targeting.xlsx")
    worksheet=workbook.sheet_by_index(0)
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    table=list()
    record=list()

    for x in range(total_rows):
        for y in range(total_columns):


            record.append(worksheet.cell(x,y).value)
        if x !=0:
            table.append(record)

        record=[]
    #print(table)
    return table
def write_adunits_ids(Adunit):
    workbook = xlsxwriter.Workbook('Adunits_IDs.xlsx')
    worksheet = workbook.add_worksheet('ADUnits')

    row = 0
    col = 0
    worksheet.write(row,col,"Adunit IDs")
    row+=1
    worksheet.write(row,col,Adunit)



def read_adunit_ids():
    workbook =xlrd.open_workbook("Adunits_IDs.xlsx")
    worksheet=workbook.sheet_by_index(0)
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    table=list()
    record=list()

    for x in range(total_rows):
        for y in range(total_columns):


            record.append(worksheet.cell(x,y).value)
        if x !=0:
            table.append(record)

        record=[]
    #print(table)
    return table

def write_campaign_ids(campaign):
    workbook = xlsxwriter.Workbook('Campaign_IDs.xlsx')
    worksheet = workbook.add_worksheet('Campaign_ID')

    row = 0
    col = 0
    worksheet.write(row,col,"Campaign IDs")
    row+=1
    worksheet.write(row,col,campaign)

def read_campaign_ids():
    workbook =xlrd.open_workbook("Campaign_IDs.xlsx")
    worksheet=workbook.sheet_by_index(0)
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    total_rows=worksheet.nrows
    total_columns=worksheet.ncols
    table=list()
    record=list()

    for x in range(total_rows):
        for y in range(total_columns):


            record.append(worksheet.cell(x,y).value)
        if x !=0:
            table.append(record)

        record=[]
    #print(table)
    return table